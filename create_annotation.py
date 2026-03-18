"""
UzLegalQA — Annotatsiya Excel fayli yaratish
Ishlatish: python create_annotation.py

Kerakli kutubxonalar:
  pip install rank_bm25 datasets openpyxl numpy
"""

import re
from datasets import load_dataset
from rank_bm25 import BM25Okapi
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# ──────────────────────────────────────────────
# SOZLAMALAR
# ──────────────────────────────────────────────

HF_DATASET_NAME = "Sohiba01/uzbek-legal-ir"
OUTPUT_EXCEL    = "UzLegalQA_Annotation.xlsx"
MIN_CHUNK_LEN   = 30
TOP_K           = 10

INTENT_COUNTS = {
    "Procedural":  15,
    "Factual":     15,
    "Consequence": 12,
    "Temporal":     8,
}

SAMPLE_QUERIES = {
    "Procedural": [
        "Ishdan bo'shatish tartibi qanday?",
        "Nikoh ro'yxatdan o'tkazish tartibi?",
        "Fuqarolik hujjatlarini olish tartibi?",
        "Soliq to'lash tartibi qanday?",
        "Mehnat shartnomasi tuzish tartibi?",
        "Meros olish tartibi qanday?",
        "Mulkni qayta ro'yxatdan o'tkazish tartibi?",
        "Da'vo ariza berish tartibi?",
        "Turar-joy sotib olish tartibi?",
        "Ijtimoiy nafaqa olish tartibi?",
        "Avtomobil ro'yxatdan o'tkazish tartibi?",
        "Tadbirkorlik litsenziyasi olish tartibi?",
        "Bank hisobini ochish tartibi?",
        "Pensiya rasmiylashtirish tartibi?",
        "Korxona tashkil etish tartibi?",
    ],
    "Factual": [
        "Mehnat shartnomasi nima?",
        "Fuqarolik huquqi nima?",
        "Soliq imtiyozi nima?",
        "Jinoiy javobgarlik nima?",
        "Voyaga yetmaganlar huquqi nima?",
        "Mulk huquqi nima?",
        "Ijara shartnomasi nima?",
        "Oilaviy huquq nima?",
        "Ma'muriy jarima nima?",
        "Sud jarayoni nima?",
        "Kafolat nima?",
        "Advokat nima?",
        "Notarius nima?",
        "Prokuratura nima?",
        "Arbitraj nima?",
    ],
    "Consequence": [
        "Shartnomani buzsa nima bo'ladi?",
        "Soliq to'lamaslik oqibati nima?",
        "Jinoyat uchun jazo nima?",
        "Mehnat huquqini buzish oqibati?",
        "Trafik qoidasini buzsa nima bo'ladi?",
        "Meros rad etilsa nima bo'ladi?",
        "Nikoh bekor qilinsa mulk qanday bo'linadi?",
        "Mulkni noqonuniy egallash oqibati?",
        "Soxta hujjat tayyorlash jazosi nima?",
        "Korrupsiya uchun jazo nima?",
        "Firibgarlik uchun jazo nima?",
        "Turar-joy qoidasini buzish oqibati?",
    ],
    "Temporal": [
        "Ta'til muddati qancha?",
        "Shartnoma muddati qancha bo'ladi?",
        "Hisobot topshirish muddati qancha?",
        "Da'vo muddati qancha?",
        "Litsenziya amal qilish muddati?",
        "Pensiya yoshi necha?",
        "Mehnat staji muddati qancha?",
        "Sud ko'rib chiqish muddati qancha?",
    ],
}

# ──────────────────────────────────────────────
# 1. DATASET YUKLASH
# ──────────────────────────────────────────────

print("Dataset yuklanmoqda...")
try:
    ds = load_dataset(HF_DATASET_NAME)
    print(f"  Splitlar: {list(ds.keys())}")
except Exception as e:
    print(f"XATO: {e}")
    exit(1)

chunks_ds = ds["train"]
print(f"  Jami: {len(chunks_ds)} ta satr")

chunk_fields = chunks_ds.column_names
text_field = next(
    (f for f in ["text", "content", "chunk_text", "article_text"] if f in chunk_fields),
    chunk_fields[0]
)
id_field = next(
    (f for f in ["chunk_id", "id", "article_id"] if f in chunk_fields),
    None
)
code_field = next(
    (f for f in ["code", "codex", "legal_code", "domain"] if f in chunk_fields),
    None
)
print(f"  Matn maydoni: '{text_field}'")

chunks = []
for i, row in enumerate(chunks_ds):
    text = (row.get(text_field) or "").strip()
    if len(text) < MIN_CHUNK_LEN:
        continue
    chunks.append({
        "text":     text,
        "chunk_id": row.get(id_field, i) if id_field else i,
        "code":     row.get(code_field, "") if code_field else "",
    })
print(f"  Filtrlashdan keyin: {len(chunks)} chunk")

# ──────────────────────────────────────────────
# 2. QUERYLAR
# ──────────────────────────────────────────────

selected_queries = []
qnum = 1
for intent, count in INTENT_COUNTS.items():
    for q_text in SAMPLE_QUERIES[intent][:count]:
        selected_queries.append({
            "query_id": f"Q{qnum:03d}",
            "text":     q_text,
            "intent":   intent,
        })
        qnum += 1
print(f"\n{len(selected_queries)} ta query tanlandi")

# ──────────────────────────────────────────────
# 3. BM25
# ──────────────────────────────────────────────

def split_concat_words(text):
    text = re.sub(r"([a-zʼʻ])([A-Z])", r"\1 \2", text)
    text = re.sub(r"([a-zA-Z])(\d)", r"\1 \2", text)
    text = re.sub(r"(\d)([a-zA-Z])", r"\1 \2", text)
    return text

def tokenize(text):
    text = split_concat_words(text)
    text = re.sub(r"[^\w\s]", " ", text.lower())
    return [t for t in text.split() if len(t) > 1]

print("\nBM25 indeks qurilmoqda...")
corpus_tokens = [tokenize(c["text"]) for c in chunks]
bm25 = BM25Okapi(corpus_tokens, k1=1.5, b=0.75)
print(f"  Tayyor: {len(corpus_tokens)} chunk")

# ──────────────────────────────────────────────
# 4. TOP-10 QIDIRISH
# ──────────────────────────────────────────────

print("\nTop-10 qidirilmoqda...")
results = []
for q in selected_queries:
    q_tokens = tokenize(q["text"])
    scores   = bm25.get_scores(q_tokens)
    top_idx  = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)[:TOP_K]
    top_chunks = []
    for rank, idx in enumerate(top_idx, 1):
        raw_text = split_concat_words(chunks[idx]["text"])
        display  = f"[{chunks[idx]['code']}] {raw_text}" if chunks[idx]["code"] else raw_text
        top_chunks.append({
            "rank":     rank,
            "chunk_id": str(chunks[idx]["chunk_id"]),
            "text":     display[:600] + ("..." if len(display) > 600 else ""),
            "score":    round(float(scores[idx]), 4),
        })
    results.append({"query": q, "top10": top_chunks})
print(f"  {len(results)} ta query uchun top-10 tayyor")

# ──────────────────────────────────────────────
# 5. EXCEL YARATISH
# ──────────────────────────────────────────────

print(f"\nExcel yaratilmoqda: {OUTPUT_EXCEL}")
wb = Workbook()

# ── helpers ──────────────────────────────────

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

INTENT_COLORS = {
    "Procedural":  "E8F5E9",
    "Factual":     "E3F2FD",
    "Consequence": "FFF8E1",
    "Temporal":    "FCE4EC",
}

INTENT_HEADER_COLORS = {
    "Procedural":  "2E7D32",
    "Factual":     "1565C0",
    "Consequence": "F57F17",
    "Temporal":    "AD1457",
}

# ══════════════════════════════════════════════
# SHEET 1: KO'RSATMALAR
# ══════════════════════════════════════════════

ws1 = wb.active
ws1.title = "Ko'rsatmalar"

ws1.column_dimensions["A"].width = 15
ws1.column_dimensions["B"].width = 80

title_font  = Font(name="Arial", size=14, bold=True, color="1A237E")
head2_font  = Font(name="Arial", size=11, bold=True, color="333333")
body_font   = Font(name="Arial", size=10, color="222222")
note_font   = Font(name="Arial", size=10, italic=True, color="555555")

def ws1_row(ws, row, label, value, lf=None, vf=None):
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = lf or head2_font
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2 = ws.cell(row=row, column=2, value=value)
    c2.font = vf or body_font
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 30

r = 1
c = ws1.cell(row=r, column=1,
             value="UzLegalQA — Annotatsiya Ko'rsatmalari")
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center")
ws1.merge_cells(f"A{r}:B{r}")
ws1.row_dimensions[r].height = 36
r += 1

instructions = [
    ("MAQSAD",
     "Ushbu fayl BM25 tizimi tomonidan topilgan matn parchalarining "
     "qanchalik tegishli ekanligini baholash uchun mo'ljallangan. "
     "Har bir so'rov (query) uchun 10 ta matn parchasi berilgan. "
     "Siz ularni 0, 1 yoki 2 ball bilan baholashingiz kerak."),
    ("BAL SHKALASI", ""),
    ("  0 — Tegishli emas",
     "Matn parcha so'rovga hech qanday aloqasi yo'q. "
     "Boshqa mavzu yoki umuman boshqa soha haqida."),
    ("  1 — Qisman tegishli",
     "Matn parcha so'rovga aloqador bo'lgan ba'zi ma'lumotlarni o'z ichiga oladi, "
     "lekin to'liq javob bermaydi."),
    ("  2 — To'liq tegishli",
     "Matn parcha so'rovga aniq va to'liq javob beradi yoki "
     "asosiy ma'lumotni o'z ichiga oladi."),
    ("QANDAY TO'LDIRISH",
     "1. 'Annotatsiya' varaqiga o'ting.\n"
     "2. Sariq fonda 'Bal (0/1/2)' ustuniga faqat 0, 1 yoki 2 kiriting.\n"
     "3. 'Izoh' ustuniga ixtiyoriy izoh yozishingiz mumkin.\n"
     "4. Boshqa ustunlarni o'zgartirmang."),
    ("MUHIM ESLATMALAR",
     "• Har bir so'rov uchun kamida 1 ta parcha 2 ball olishi kerak (agar mavjud bo'lsa).\n"
     "• Baholash mustaqil amalga oshirilsin — boshqalar bilan maslahatlashmang.\n"
     "• Shubha bo'lsa, 1 ball bering."),
    ("RANGLAR",
     "Ko'k fon = Factual so'rovlar\n"
     "Yashil fon = Procedural so'rovlar\n"
     "Sariq fon = Consequence so'rovlar\n"
     "Pushti fon = Temporal so'rovlar"),
    ("JAMI",
     f"50 ta so'rov × 10 ta parcha = 500 ta qator baholanadi."),
]

for label, value in instructions:
    ws1_row(ws1, r, label, value)
    r += 1

# ══════════════════════════════════════════════
# SHEET 2: ANNOTATSIYA
# ══════════════════════════════════════════════

ws2 = wb.create_sheet("Annotatsiya")

# Ustun kengliklarini sozlash
col_widths = [5, 9, 14, 38, 6, 80, 12, 10, 12, 25]
col_names  = ["#", "Query_ID", "Intent", "Query matni", "Rank",
              "Matn parchasi (birinchi 600 belgi)", "Chunk_ID",
              "BM25 ball", "Bal (0/1/2)", "Izoh"]
for i, (w, name) in enumerate(zip(col_widths, col_names), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Sarlavha satri
hdr_fill = header_fill("1A237E")
hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
hdr_border = thin_border()
for col, name in enumerate(col_names, 1):
    c = ws2.cell(row=1, column=col, value=name)
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = hdr_border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[1].height = 32
ws2.freeze_panes = "A2"

row = 2
for res_idx, res in enumerate(results):
    q     = res["query"]
    top10 = res["top10"]
    bg    = INTENT_COLORS.get(q["intent"], "F5F5F5")
    r_fill = cell_fill(bg)
    y_fill = cell_fill("FFFDE7")
    bd     = thin_border()

    for rank_idx, chunk in enumerate(top10):
        is_first = rank_idx == 0

        def wc(col, val, fill=None, font=None, align=None):
            c = ws2.cell(row=row, column=col, value=val)
            c.fill = fill or r_fill
            c.border = bd
            c.font = font or Font(name="Arial", size=9, color="222222")
            c.alignment = align or Alignment(vertical="center", wrap_text=True)
            return c

        # Col 1: #
        wc(1, (res_idx + 1) if is_first else "",
           font=Font(name="Arial", size=10, bold=True, color="333333"),
           align=Alignment(horizontal="center", vertical="center"))

        # Col 2: Query_ID
        wc(2, q["query_id"] if is_first else "",
           font=Font(name="Arial", size=9, bold=True, color="1F4E79"),
           align=Alignment(horizontal="center", vertical="center"))

        # Col 3: Intent
        intent_color = INTENT_HEADER_COLORS.get(q["intent"], "333333")
        wc(3, q["intent"] if is_first else "",
           font=Font(name="Arial", size=9, bold=True, color=intent_color),
           align=Alignment(horizontal="center", vertical="center"))

        # Col 4: Query text
        wc(4, q["text"] if is_first else "",
           align=Alignment(horizontal="left", vertical="center", wrap_text=True))

        # Col 5: Rank
        wc(5, chunk["rank"],
           font=Font(name="Arial", size=9, bold=True, color="555555"),
           align=Alignment(horizontal="center", vertical="center"))

        # Col 6: Chunk text
        wc(6, chunk["text"],
           align=Alignment(horizontal="left", vertical="center", wrap_text=True))

        # Col 7: Chunk_ID
        wc(7, chunk["chunk_id"],
           font=Font(name="Arial", size=9, color="777777"),
           align=Alignment(horizontal="center", vertical="center"))

        # Col 8: BM25 score
        wc(8, chunk["score"],
           font=Font(name="Arial", size=9, color="555555"),
           align=Alignment(horizontal="center", vertical="center"))

        # Col 9: Bal — sariq, bo'sh
        c9 = ws2.cell(row=row, column=9, value="")
        c9.fill = y_fill
        c9.border = bd
        c9.font = Font(name="Arial", size=12, bold=True, color="B71C1C")
        c9.alignment = Alignment(horizontal="center", vertical="center")

        # Col 10: Izoh
        c10 = ws2.cell(row=row, column=10, value="")
        c10.fill = r_fill
        c10.border = bd
        c10.font = Font(name="Arial", size=9, color="444444")
        c10.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws2.row_dimensions[row].height = 80
        row += 1

    # Separator satri
    sep_fill = cell_fill("BDBDBD")
    for col in range(1, 11):
        c = ws2.cell(row=row, column=col, value="")
        c.fill = sep_fill
    ws2.row_dimensions[row].height = 3
    row += 1

# ══════════════════════════════════════════════
# SHEET 3: METRIKALAR JADVALI
# ══════════════════════════════════════════════

ws3 = wb.create_sheet("Metrikalar")

# Annotation sheet column reference for Bal (column I = 9)
# We need to map query+rank → Excel row in ws2

# Build row_map: (query_index, rank) → row number in Annotatsiya sheet
# row 1 is header, then 10 data rows + 1 sep per query
def annot_row(q_idx, rank_idx):
    """Row number in Annotatsiya sheet for query q_idx (0-based), rank rank_idx (0-based)."""
    return 2 + q_idx * 11 + rank_idx  # 10 rows + 1 separator per query

# Metrics sheet columns
m_col_widths = [6, 10, 14, 40, 10, 10, 10, 10]
m_col_names  = ["#", "Query_ID", "Intent", "Query matni",
                "MRR", "P@5", "R@10", "nDCG@10"]
for i, (w, name) in enumerate(zip(m_col_widths, m_col_names), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

hdr2_fill = header_fill("37474F")
for col, name in enumerate(m_col_names, 1):
    c = ws3.cell(row=1, column=col, value=name)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = hdr2_fill
    c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[1].height = 28
ws3.freeze_panes = "A2"

# For each query, add a row with formula-based metric placeholders
# (Actual metrics are computed by calculate_metrics.py; here we show BM25 scores as proxy)
# We add a note explaining this sheet is for reference after annotation.

note_row = 2
note_cell = ws3.cell(row=note_row, column=1,
    value="ESLATMA: Bu jadval calculate_metrics.py skripti tomonidan to'ldiriladi. "
          "Annotatsiyadan keyin skriptni ishga tushiring.")
note_cell.font = Font(name="Arial", size=10, italic=True, color="B71C1C")
note_cell.alignment = Alignment(wrap_text=True)
ws3.merge_cells(f"A{note_row}:H{note_row}")
ws3.row_dimensions[note_row].height = 32
mr = note_row + 1

intent_order = ["Procedural", "Factual", "Consequence", "Temporal"]
intent_start_rows = {}  # intent → list of data rows in ws3

q_idx = 0
for intent in intent_order:
    # Intent header
    ih_fill = header_fill(INTENT_HEADER_COLORS[intent])
    c = ws3.cell(row=mr, column=1, value=f"── {intent} ──")
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = ih_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.merge_cells(f"A{mr}:H{mr}")
    ws3.row_dimensions[mr].height = 20
    mr += 1

    intent_data_rows = []
    for q in selected_queries:
        if q["intent"] != intent:
            continue
        # Placeholder row — calculate_metrics.py will overwrite
        bg = INTENT_COLORS[intent]
        rf = cell_fill(bg)
        bd = thin_border()
        cells_vals = [q_idx + 1, q["query_id"], q["intent"], q["text"],
                      "", "", "", ""]
        for col, val in enumerate(cells_vals, 1):
            c = ws3.cell(row=mr, column=col, value=val)
            c.fill = rf
            c.border = bd
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(col == 4))
        ws3.row_dimensions[mr].height = 22
        intent_data_rows.append(mr)
        q_idx += 1
        mr += 1

    intent_start_rows[intent] = intent_data_rows

# Intent averages
mr_avg_start = {}
for intent in intent_order:
    rows = intent_start_rows[intent]
    if not rows:
        continue
    ih_fill = header_fill("ECEFF1")
    c = ws3.cell(row=mr, column=1, value=f"{intent} o'rtacha")
    c.font = Font(name="Arial", size=9, bold=True, color="333333")
    c.fill = ih_fill
    ws3.merge_cells(f"A{mr}:D{mr}")
    bd = thin_border()
    for col in range(1, 5):
        ws3.cell(row=mr, column=col).border = bd
    for col_idx, col_letter in enumerate(["E", "F", "G", "H"], 0):
        row_refs = ",".join(f"{col_letter}{r}" for r in rows)
        formula = f"=IFERROR(AVERAGE({row_refs}),\"\")"
        c = ws3.cell(row=mr, column=5 + col_idx, value=formula)
        c.font = Font(name="Arial", size=9, bold=True, color="1A237E")
        c.fill = ih_fill
        c.border = bd
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[mr].height = 22
    mr += 1

# Overall average
overall_avg_row = mr
oa_fill = header_fill("263238")
c = ws3.cell(row=mr, column=1, value="UMUMIY O'RTACHA")
c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
c.fill = oa_fill
ws3.merge_cells(f"A{mr}:D{mr}")
bd = thin_border()
for col in range(1, 5):
    ws3.cell(row=mr, column=col).border = bd
all_data_rows = [r for rows in intent_start_rows.values() for r in rows]
for col_idx, col_letter in enumerate(["E", "F", "G", "H"], 0):
    row_refs = ",".join(f"{col_letter}{r}" for r in all_data_rows)
    formula = f"=IFERROR(AVERAGE({row_refs}),\"\")"
    c = ws3.cell(row=mr, column=5 + col_idx, value=formula)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = oa_fill
    c.border = bd
    c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[mr].height = 26

# ── Saqlash ──────────────────────────────────

wb.save(OUTPUT_EXCEL)
print(f"\n✅ Tayyor! '{OUTPUT_EXCEL}' yaratildi.")
print(f"   • Ko'rsatmalar  — 1-varaq")
print(f"   • Annotatsiya   — 2-varaq  ({len(selected_queries)} so'rov × 10 = {len(selected_queries)*10} qator)")
print(f"   • Metrikalar    — 3-varaq  (calculate_metrics.py ishlatgandan keyin to'ladi)")
print(f"\n   Sariq ustun 'Bal (0/1/2)' — siz to'ldirasiz (0, 1 yoki 2)")
