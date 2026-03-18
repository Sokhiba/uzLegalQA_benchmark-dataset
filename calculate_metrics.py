"""
UzLegalQA — IR metrikalarini hisoblash
Ishlatish: python calculate_metrics.py

Kerakli kutubxonalar:
  pip install openpyxl numpy
"""

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

INPUT_EXCEL  = "UzLegalQA_Annotation.xlsx"
OUTPUT_EXCEL = "UzLegalQA_Metrics.xlsx"

ANNOTATION_SHEET = "Annotatsiya"
METRICS_SHEET    = "Metrikalar"

TOP_K       = 10
REL_THRESH  = 1   # Bal >= 1 → relevant (for binary metrics)

INTENT_ORDER = ["Procedural", "Factual", "Consequence", "Temporal"]

INTENT_COLORS = {
    "Procedural":  "E8F5E9",
    "Factual":     "E3F2FD",
    "Consequence": "FFF8E1",
    "Temporal":    "FCE4EC",
}
INTENT_HEADER_COLORS = {
    "Procedural":  "2E7D32",
    "Factual":     "1565C0",
    "Consequence": "F57F17",
    "Temporal":    "AD1457",
}

# ──────────────────────────────────────────────
# 1. ANNOTATSIYANI O'QISH
# ──────────────────────────────────────────────

print(f"Fayl o'qilmoqda: {INPUT_EXCEL}")

try:
    wb_in = load_workbook(INPUT_EXCEL, data_only=True)
except FileNotFoundError:
    print(f"XATO: '{INPUT_EXCEL}' topilmadi. Avval create_annotation.py ishga tushiring.")
    exit(1)

if ANNOTATION_SHEET not in wb_in.sheetnames:
    print(f"XATO: '{ANNOTATION_SHEET}' varag'i topilmadi.")
    exit(1)

ws = wb_in[ANNOTATION_SHEET]

# Sarlavha satri (row 1) o'tkazib yuboriladi.
# Har bir query — 10 data satri + 1 separator satri (jami 11 satr blok).
# Ustunlar:
#   A=1:#  B=2:Query_ID  C=3:Intent  D=4:Query text
#   E=5:Rank  F=6:Chunk text  G=7:Chunk_ID  H=8:BM25 score
#   I=9:Bal (0/1/2)  J=10:Izoh

# Col indices (1-based)
COL_NUM      = 1
COL_QID      = 2
COL_INTENT   = 3
COL_QTEXT    = 4
COL_RANK     = 5
COL_CHUNK    = 6
COL_CHUNKID  = 7
COL_BM25     = 8
COL_BAL      = 9
COL_NOTE     = 10

queries = []
current_query = None

max_row = ws.max_row
data_row = 2  # skip header

while data_row <= max_row:
    row_vals = [ws.cell(row=data_row, column=c).value for c in range(1, 11)]
    num_val  = row_vals[COL_NUM - 1]
    bal_val  = row_vals[COL_BAL - 1]
    rank_val = row_vals[COL_RANK - 1]

    # Separator satri: barcha qiymatlar bo'sh yoki None
    if all(v is None or str(v).strip() == "" for v in row_vals):
        data_row += 1
        continue

    # Yangi query bloki boshlansa
    if num_val is not None and str(num_val).strip() not in ("", "None"):
        current_query = {
            "num":    int(num_val),
            "id":     str(row_vals[COL_QID - 1] or ""),
            "intent": str(row_vals[COL_INTENT - 1] or "Unknown"),
            "text":   str(row_vals[COL_QTEXT - 1] or ""),
            "grades": [],
            "scores": [],
        }
        queries.append(current_query)

    if current_query is not None and rank_val is not None:
        # Bal
        try:
            grade = int(float(str(bal_val))) if bal_val is not None and str(bal_val).strip() != "" else 0
        except (ValueError, TypeError):
            grade = 0
        grade = max(0, min(2, grade))

        # BM25 score
        try:
            bm25_score = float(row_vals[COL_BM25 - 1])
        except (TypeError, ValueError):
            bm25_score = 0.0

        current_query["grades"].append(grade)
        current_query["scores"].append(bm25_score)

    data_row += 1

print(f"  {len(queries)} ta query o'qildi")

# Tekshirish
missing_grades = [q["id"] for q in queries if len(q["grades"]) == 0]
incomplete     = [q["id"] for q in queries if len(q["grades"]) < TOP_K]
unfilled       = [q["id"] for q in queries if all(g == 0 for g in q["grades"])]

if missing_grades:
    print(f"  OGOHLANTIRISH: {len(missing_grades)} ta query uchun bal topilmadi: {missing_grades[:5]}")
if incomplete:
    print(f"  OGOHLANTIRISH: {len(incomplete)} ta query to'liq emas (<{TOP_K} bal)")
if unfilled:
    print(f"  OGOHLANTIRISH: {len(unfilled)} ta query uchun barcha ballar 0 (to'ldirilmagan?)")

# ──────────────────────────────────────────────
# 2. METRIKALAR HISOBLASH
# ──────────────────────────────────────────────

def dcg(grades, k):
    """Discounted Cumulative Gain at k."""
    g = grades[:k]
    return sum(rel / np.log2(idx + 2) for idx, rel in enumerate(g))

def ndcg(grades, k):
    """Normalized DCG at k."""
    ideal = sorted(grades, reverse=True)
    idcg  = dcg(ideal, k)
    if idcg == 0:
        return 0.0
    return dcg(grades, k) / idcg

def mrr(grades):
    """Mean Reciprocal Rank (first grade >= REL_THRESH)."""
    for i, g in enumerate(grades):
        if g >= REL_THRESH:
            return 1.0 / (i + 1)
    return 0.0

def precision_at_k(grades, k):
    """Precision@k (binary relevance)."""
    g = grades[:k]
    rel = sum(1 for x in g if x >= REL_THRESH)
    return rel / k

def recall_at_k(grades, k):
    """Recall@k — fraction of all relevant docs retrieved in top-k."""
    total_rel = sum(1 for x in grades if x >= REL_THRESH)
    if total_rel == 0:
        return 0.0
    retrieved_rel = sum(1 for x in grades[:k] if x >= REL_THRESH)
    return retrieved_rel / total_rel

def average_precision(grades):
    """Average Precision (AP) for MAP."""
    total_rel = sum(1 for x in grades if x >= REL_THRESH)
    if total_rel == 0:
        return 0.0
    ap = 0.0
    running_rel = 0
    for i, g in enumerate(grades):
        if g >= REL_THRESH:
            running_rel += 1
            ap += running_rel / (i + 1)
    return ap / total_rel

query_metrics = []
for q in queries:
    grades = q["grades"]
    if not grades:
        grades = [0] * TOP_K

    m = {
        "num":    q["num"],
        "id":     q["id"],
        "intent": q["intent"],
        "text":   q["text"],
        "ndcg10": round(ndcg(grades, TOP_K), 4),
        "mrr":    round(mrr(grades), 4),
        "map":    round(average_precision(grades), 4),
        "r10":    round(recall_at_k(grades, TOP_K), 4),
        "p5":     round(precision_at_k(grades, 5), 4),
        "grades": grades,
    }
    query_metrics.append(m)

# ──────────────────────────────────────────────
# 3. NATIJALARI CHIQARISH
# ──────────────────────────────────────────────

print("\n" + "═" * 72)
print(f"{'ID':<8} {'Intent':<14} {'nDCG@10':>8} {'MRR':>8} {'MAP':>8} {'R@10':>8} {'P@5':>8}")
print("═" * 72)

intent_groups = defaultdict(list)
for m in query_metrics:
    intent_groups[m["intent"]].append(m)

for intent in INTENT_ORDER:
    group = intent_groups.get(intent, [])
    if not group:
        continue
    print(f"\n  [{intent}]")
    for m in group:
        print(f"  {m['id']:<8} {m['intent']:<14} "
              f"{m['ndcg10']:>8.4f} {m['mrr']:>8.4f} "
              f"{m['map']:>8.4f} {m['r10']:>8.4f} {m['p5']:>8.4f}")
    g_ndcg = np.mean([m["ndcg10"] for m in group])
    g_mrr  = np.mean([m["mrr"]    for m in group])
    g_map  = np.mean([m["map"]    for m in group])
    g_r10  = np.mean([m["r10"]    for m in group])
    g_p5   = np.mean([m["p5"]     for m in group])
    print(f"  {'─'*68}")
    print(f"  {'AVG':<8} {intent:<14} "
          f"{g_ndcg:>8.4f} {g_mrr:>8.4f} "
          f"{g_map:>8.4f} {g_r10:>8.4f} {g_p5:>8.4f}")

all_ndcg = np.mean([m["ndcg10"] for m in query_metrics])
all_mrr  = np.mean([m["mrr"]    for m in query_metrics])
all_map  = np.mean([m["map"]    for m in query_metrics])
all_r10  = np.mean([m["r10"]    for m in query_metrics])
all_p5   = np.mean([m["p5"]     for m in query_metrics])

print("\n" + "═" * 72)
print(f"  {'OVERALL':<8} {'ALL':<14} "
      f"{all_ndcg:>8.4f} {all_mrr:>8.4f} "
      f"{all_map:>8.4f} {all_r10:>8.4f} {all_p5:>8.4f}")
print("═" * 72)

# ──────────────────────────────────────────────
# 4. METRICS SHEETINI YANGILASH
# ──────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# Update existing annotation workbook's Metrikalar sheet
wb_out = load_workbook(INPUT_EXCEL)

if METRICS_SHEET in wb_out.sheetnames:
    del wb_out[METRICS_SHEET]

ws3 = wb_out.create_sheet(METRICS_SHEET)

# Column setup
col_widths = [6, 10, 14, 38, 10, 10, 10, 10, 10]
col_names  = ["#", "Query_ID", "Intent", "Query matni",
              "nDCG@10", "MRR", "MAP", "R@10", "P@5"]
for i, (w, name) in enumerate(zip(col_widths, col_names), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Header
hdr_fill = cell_fill("263238")
for col, name in enumerate(col_names, 1):
    c = ws3.cell(row=1, column=col, value=name)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = hdr_fill
    c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[1].height = 28
ws3.freeze_panes = "A2"

mr = 2
intent_data_rows = defaultdict(list)

for intent in INTENT_ORDER:
    group = intent_groups.get(intent, [])
    if not group:
        continue

    # Intent header row
    ih_fill = cell_fill(INTENT_HEADER_COLORS[intent])
    c = ws3.cell(row=mr, column=1, value=f"── {intent} ──")
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = ih_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.merge_cells(f"A{mr}:I{mr}")
    for col in range(1, 10):
        ws3.cell(row=mr, column=col).border = thin_border()
    ws3.row_dimensions[mr].height = 20
    mr += 1

    bg    = INTENT_COLORS[intent]
    rf    = cell_fill(bg)
    bd    = thin_border()

    for m in group:
        vals = [m["num"], m["id"], m["intent"], m["text"],
                m["ndcg10"], m["mrr"], m["map"], m["r10"], m["p5"]]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=mr, column=col, value=val)
            c.fill = rf
            c.border = bd
            c.font = Font(name="Arial", size=9)
            if col <= 4:
                c.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = "0.0000"
        ws3.row_dimensions[mr].height = 22
        intent_data_rows[intent].append(mr)
        mr += 1

    # Intent average
    rows = intent_data_rows[intent]
    avg_fill = cell_fill("ECEFF1")
    c = ws3.cell(row=mr, column=1, value=f"{intent} o'rtacha")
    c.font = Font(name="Arial", size=9, bold=True, color="333333")
    c.fill = avg_fill
    ws3.merge_cells(f"A{mr}:D{mr}")
    for col in range(1, 5):
        ws3.cell(row=mr, column=col).fill = avg_fill
        ws3.cell(row=mr, column=col).border = bd
    metrics_cols = ["E", "F", "G", "H", "I"]
    metric_vals  = [
        np.mean([m["ndcg10"] for m in group]),
        np.mean([m["mrr"]    for m in group]),
        np.mean([m["map"]    for m in group]),
        np.mean([m["r10"]    for m in group]),
        np.mean([m["p5"]     for m in group]),
    ]
    for col_idx, (col_letter, val) in enumerate(zip(metrics_cols, metric_vals)):
        c = ws3.cell(row=mr, column=5 + col_idx, value=round(val, 4))
        c.font = Font(name="Arial", size=9, bold=True, color="1A237E")
        c.fill = avg_fill
        c.border = bd
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.0000"
    ws3.row_dimensions[mr].height = 22
    mr += 1

# Overall average
oa_fill = cell_fill("1A237E")
c = ws3.cell(row=mr, column=1, value="UMUMIY O'RTACHA")
c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
c.fill = oa_fill
ws3.merge_cells(f"A{mr}:D{mr}")
for col in range(1, 5):
    ws3.cell(row=mr, column=col).fill = oa_fill
    ws3.cell(row=mr, column=col).border = bd
overall_vals = [all_ndcg, all_mrr, all_map, all_r10, all_p5]
for col_idx, val in enumerate(overall_vals):
    c = ws3.cell(row=mr, column=5 + col_idx, value=round(val, 4))
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = oa_fill
    c.border = bd
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.0000"
ws3.row_dimensions[mr].height = 28

wb_out.save(OUTPUT_EXCEL)

print(f"\n✅ Metrikalar saqlandi: '{OUTPUT_EXCEL}'")
print(f"   'Metrikalar' varag'ida barcha natijalar mavjud.")

# ──────────────────────────────────────────────
# 5. QISQACHA XULOSA (maqola uchun)
# ──────────────────────────────────────────────

print("\n" + "═" * 50)
print("  MAQOLA UCHUN QISQACHA XULOSA (BM25 baseline)")
print("═" * 50)
print(f"  nDCG@10 : {all_ndcg:.4f}")
print(f"  MRR     : {all_mrr:.4f}")
print(f"  MAP     : {all_map:.4f}")
print(f"  R@10    : {all_r10:.4f}")
print(f"  P@5     : {all_p5:.4f}")
print("═" * 50)
print(f"  Baholangan querylar soni : {len(query_metrics)}")
print(f"  Baholangan parchalar soni: {sum(len(q['grades']) for q in queries)}")
